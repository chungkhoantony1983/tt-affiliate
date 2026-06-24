#!/usr/bin/env python3
"""
顧客共有用資料のエクスポートスクリプト
- CSV → Excel変換
- Markdown → PDF変換
"""

import csv
import sys
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import markdown
from weasyprint import HTML, CSS

def csv_to_excel(csv_path: Path, excel_path: Path):
    """CSVファイルをExcelに変換（スタイル付き）"""
    print(f"Converting CSV to Excel: {csv_path} -> {excel_path}")

    # CSVを読み込み
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        rows = list(reader)

    # Excelワークブック作成
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "課題管理表"

    # スタイル定義
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # データ書き込み
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = cell_alignment

            # ヘッダー行のスタイル
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

    # 列幅調整
    column_widths = {
        'A': 12,  # ID
        'B': 35,  # 課題名
        'C': 10,  # 優先度
        'D': 12,  # ステータス
        'E': 20,  # 担当
        'F': 12,  # 期限
        'G': 50,  # 説明
        'H': 30,  # 関連ドキュメント
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 行の高さ調整（自動調整）
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = None  # 自動調整

    # フリーズペイン（ヘッダー固定）
    ws.freeze_panes = 'A2'

    # フィルター設定
    ws.auto_filter.ref = ws.dimensions

    # 保存
    wb.save(excel_path)
    print(f"✅ Excel saved: {excel_path}")


def markdown_to_pdf(md_path: Path, pdf_path: Path):
    """MarkdownファイルをPDFに変換"""
    print(f"Converting Markdown to PDF: {md_path} -> {pdf_path}")

    # Markdownを読み込み
    with open(md_path, 'r', encoding='utf-8') as f:
        md_content = f.read()

    # MarkdownをHTMLに変換
    html_content = markdown.markdown(
        md_content,
        extensions=['tables', 'fenced_code', 'codehilite', 'nl2br']
    )

    # HTMLテンプレート
    html_template = f"""
    <!DOCTYPE html>
    <html lang="ja">
    <head>
        <meta charset="UTF-8">
        <title>77B-901-課題詳細</title>
        <style>
            @page {{
                size: A4;
                margin: 2cm;
            }}
            body {{
                font-family: "Noto Sans JP", "Hiragino Sans", "MS Gothic", sans-serif;
                font-size: 10pt;
                line-height: 1.6;
                color: #333;
            }}
            h1 {{
                font-size: 18pt;
                border-bottom: 2px solid #4472C4;
                padding-bottom: 0.5em;
                margin-top: 1.5em;
                page-break-before: auto;
            }}
            h2 {{
                font-size: 16pt;
                border-bottom: 1px solid #4472C4;
                padding-bottom: 0.3em;
                margin-top: 1.2em;
                page-break-after: avoid;
            }}
            h3 {{
                font-size: 14pt;
                margin-top: 1em;
                page-break-after: avoid;
            }}
            h4 {{
                font-size: 12pt;
                margin-top: 0.8em;
            }}
            table {{
                border-collapse: collapse;
                width: 100%;
                margin: 1em 0;
                font-size: 9pt;
            }}
            th, td {{
                border: 1px solid #ccc;
                padding: 0.5em;
                text-align: left;
            }}
            th {{
                background-color: #4472C4;
                color: white;
                font-weight: bold;
            }}
            code {{
                background-color: #f5f5f5;
                padding: 0.2em 0.4em;
                border-radius: 3px;
                font-family: monospace;
            }}
            ul, ol {{
                margin: 0.5em 0;
                padding-left: 2em;
            }}
            li {{
                margin: 0.3em 0;
            }}
            strong {{
                font-weight: bold;
            }}
            hr {{
                border: none;
                border-top: 1px solid #ccc;
                margin: 1.5em 0;
            }}
        </style>
    </head>
    <body>
        {html_content}
    </body>
    </html>
    """

    # HTMLをPDFに変換
    HTML(string=html_template).write_pdf(pdf_path)
    print(f"✅ PDF saved: {pdf_path}")


def main():
    # プロジェクトルート
    project_root = Path(__file__).resolve().parent.parent.parent

    # 入力ファイル
    csv_source = project_root / "77/9_Decision_Records/77B-901-課題管理表.csv"
    md_source = project_root / "77/9_Decision_Records/77B-901-課題詳細.md"

    # 出力先（成果物フォルダ）
    today = datetime.now().strftime("%Y-%m-%d")
    deliverable_dir = project_root / "77/docs/exports"
    deliverable_dir.mkdir(parents=True, exist_ok=True)

    excel_output = deliverable_dir / "77B-901-課題管理表.xlsx"
    pdf_output = deliverable_dir / "77B-901-課題詳細.pdf"

    # 変換実行
    try:
        csv_to_excel(csv_source, excel_output)
        markdown_to_pdf(md_source, pdf_output)
        print(f"\n✅ All deliverables exported to: {deliverable_dir}")
        print(f"   - {excel_output.name}")
        print(f"   - {pdf_output.name}")
    except Exception as e:
        print(f"❌ Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
