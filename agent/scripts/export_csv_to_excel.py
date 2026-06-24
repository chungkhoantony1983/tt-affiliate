#!/usr/bin/env python3
"""
CSV を書式付き Excel（.xlsx）に変換する共通スクリプト。

主な機能:
- ヘッダー背景色 + 白文字
- 全セル罫線
- 列幅自動調整
- オートフィルター
- 成果物ディレクトリ (`docs/exports/`) への出力
"""

from __future__ import annotations

import argparse
import csv
import sys
import unicodedata
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="CSV を書式付き Excel（.xlsx）へ変換します。"
    )
    parser.add_argument("input_csv", type=Path, help="入力 CSV ファイル")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help="出力 xlsx ファイル（省略時は入力と同名 .xlsx）",
    )
    parser.add_argument(
        "--sheet-name",
        default="Sheet1",
        help="出力シート名（既定: Sheet1）",
    )
    parser.add_argument(
        "--encoding",
        default="utf-8-sig",
        help="CSV文字コード（既定: utf-8-sig）",
    )
    parser.add_argument(
        "--delimiter",
        help="CSV区切り文字（省略時は自動判定）",
    )
    parser.add_argument(
        "--to-deliverable",
        action="store_true",
        help="`docs/exports/` へ出力する",
    )
    parser.add_argument(
        "--project-root",
        type=Path,
        help="`--to-deliverable` 時のPJルート（省略時は入力ファイルから推定）",
    )
    parser.add_argument(
        "--date",
        dest="deliverable_date",
        help="`--to-deliverable` 時の日付（YYYY-MM-DD, 省略時は当日）",
    )
    parser.add_argument(
        "--filter-column",
        help="フィルタ対象の列名（この列の値が --filter-value に一致する行のみ出力）",
    )
    parser.add_argument(
        "--filter-value",
        help="フィルタ対象の値（--filter-column と組み合わせて使用）",
    )
    parser.add_argument(
        "--drop-columns",
        help="出力から除外する列名（カンマ区切りで複数指定可）",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="変換は実行せず、出力先と設定のみ表示する",
    )
    return parser.parse_args()


def infer_project_root(input_path: Path) -> Path | None:
    required = {"0_Project", "9_Decision_Records"}
    for parent in [input_path.parent, *input_path.parents]:
        try:
            dir_names = {child.name for child in parent.iterdir() if child.is_dir()}
        except OSError:
            continue

        if required.issubset(dir_names):
            return parent
    return None


def resolve_output_xlsx(args: argparse.Namespace, input_csv: Path, *, dry_run: bool = False) -> Path:
    if args.output:
        output = args.output.resolve()
        if not dry_run:
            output.parent.mkdir(parents=True, exist_ok=True)
        return output

    if not args.to_deliverable:
        return input_csv.with_suffix(".xlsx")

    if args.deliverable_date:
        try:
            export_date = date.fromisoformat(args.deliverable_date).isoformat()
        except ValueError as exc:
            raise ValueError(
                f"--date は YYYY-MM-DD 形式で指定してください: {args.deliverable_date}"
            ) from exc
    else:
        export_date = date.today().isoformat()

    project_root = args.project_root.resolve() if args.project_root else infer_project_root(input_csv)
    if project_root is None:
        raise ValueError(
            "--to-deliverable 利用時にPJルートを推定できませんでした。"
            " `--project-root <path>` を指定してください。"
        )

    deliverable_dir = project_root / "docs" / "exports"
    if not dry_run:
        deliverable_dir.mkdir(parents=True, exist_ok=True)
    return deliverable_dir / f"{input_csv.stem}.xlsx"


def detect_dialect(sample: str) -> csv.Dialect:
    try:
        return csv.Sniffer().sniff(sample, delimiters=[",", "\t", ";", "|"])
    except csv.Error:
        return csv.excel


def read_rows(input_csv: Path, encoding: str, delimiter: str | None) -> tuple[list[list[str]], str]:
    with input_csv.open("r", encoding=encoding, newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)

        if delimiter:
            reader = csv.reader(handle, delimiter=delimiter)
            rows = list(reader)
            return rows, delimiter

        dialect = detect_dialect(sample)
        reader = csv.reader(handle, dialect)
        rows = list(reader)
        return rows, dialect.delimiter


def display_width(text: str) -> int:
    width = 0
    for ch in text:
        width += 2 if unicodedata.east_asian_width(ch) in {"F", "W", "A"} else 1
    return width


def trim_sheet_name(sheet_name: str) -> str:
    cleaned = sheet_name.strip() or "Sheet1"
    return cleaned[:31]


def write_xlsx(rows: list[list[str]], output_xlsx: Path, sheet_name: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = trim_sheet_name(sheet_name)

    header_font = Font(name="Yu Gothic", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            else:
                cell.alignment = cell_alignment

    if rows:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        col_count = max(len(r) for r in rows)
        for col_idx in range(1, col_count + 1):
            values = []
            for row in rows:
                values.append(row[col_idx - 1] if col_idx - 1 < len(row) else "")
            max_width = max(display_width(str(v)) for v in values) if values else 8
            adjusted = min(max(8, max_width + 2), 60)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)


def main() -> int:
    args = parse_args()

    input_csv = args.input_csv.resolve()
    if not input_csv.exists():
        print(f"ERROR: 入力ファイルが見つかりません: {input_csv}", file=sys.stderr)
        return 1

    try:
        output_xlsx = resolve_output_xlsx(args, input_csv, dry_run=args.dry_run)
    except ValueError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    filter_column = args.filter_column
    filter_value = args.filter_value
    drop_columns = [c.strip() for c in args.drop_columns.split(",")] if args.drop_columns else []

    if args.dry_run:
        print("INFO: Dry run mode: 変換は実行しません。")
        print(f"   input : {input_csv}")
        print(f"   output: {output_xlsx}")
        print(f"   sheet : {trim_sheet_name(args.sheet_name)}")
        print(f"   enc   : {args.encoding}")
        print(f"   delim : {args.delimiter or 'auto'}")
        if filter_column:
            print(f"   filter: {filter_column} = {filter_value}")
        if drop_columns:
            print(f"   drop  : {', '.join(drop_columns)}")
        return 0

    try:
        rows, used_delim = read_rows(input_csv, args.encoding, args.delimiter)
    except UnicodeDecodeError:
        print(
            f"ERROR: 文字コード {args.encoding} で読み込めませんでした。"
            " `--encoding` を指定し直してください。",
            file=sys.stderr,
        )
        return 1
    except Exception as exc:
        print(f"ERROR: CSV読み込みに失敗しました: {exc}", file=sys.stderr)
        return 1

    if not rows:
        print("ERROR: CSVが空です。変換を中止します。", file=sys.stderr)
        return 1

    # ── 行フィルタ: 指定列の値が一致する行のみ残す ──
    original_count = len(rows) - 1  # ヘッダー除く
    if filter_column and rows:
        header = rows[0]
        if filter_column not in header:
            print(
                f"ERROR: フィルタ列 '{filter_column}' がCSVヘッダーに見つかりません。"
                f" 利用可能な列: {', '.join(header)}",
                file=sys.stderr,
            )
            return 1
        col_idx = header.index(filter_column)
        filtered = [rows[0]]  # ヘッダーは常に残す
        for row in rows[1:]:
            if col_idx < len(row) and row[col_idx] == filter_value:
                filtered.append(row)
        rows = filtered
        print(f"   Filter: {filter_column}={filter_value} → {len(rows)-1}/{original_count} 行")

    # ── 列ドロップ: 指定列を出力から除外 ──
    if drop_columns and rows:
        header = rows[0]
        drop_idxs = sorted(
            [header.index(c) for c in drop_columns if c in header],
            reverse=True,
        )
        for row in rows:
            for idx in drop_idxs:
                if idx < len(row):
                    del row[idx]
        if drop_idxs:
            print(f"   Drop  : {', '.join(drop_columns)} ({len(drop_idxs)} 列除外)")

    try:
        write_xlsx(rows, output_xlsx, args.sheet_name)
    except Exception as exc:
        print(f"ERROR: Excel書き込みに失敗しました: {exc}", file=sys.stderr)
        return 1

    print(f"OK: Excel 出力完了: {output_xlsx}")
    print(f"   Rows : {len(rows)}")
    print(f"   Delim: {used_delim!r}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
