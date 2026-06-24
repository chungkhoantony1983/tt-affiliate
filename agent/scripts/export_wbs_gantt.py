#!/usr/bin/env python3
"""
WBS CSV からガントチャート Excel を生成する共通スクリプト。

主な機能:
- フェーズ×工程（または Gate×テーマ）ごとにバーを集約描画
- 工程別色分け
- 週単位の列でバーを色付きセルとして表現
- 列名自動検出（Idemitsu / Hokkaido 両形式対応）
"""

from __future__ import annotations

import argparse
import csv
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 定数
# ---------------------------------------------------------------------------

# 列名候補マッピング（最初に見つかったものを使用）
COLUMN_CANDIDATES: dict[str, list[str]] = {
    "start_date": ["開始日", "開始"],
    "end_date": ["終了日", "終了"],
    "task_name": ["タスク名", "タスク"],
    "group": ["フェーズ", "Gate"],
    "subgroup": ["工程"],
    "stream": ["ストリーム", "Stream"],
}

# 工程の色定義 (RRGGBB without #)
PROCESS_COLORS: dict[str, str] = {
    "要件定義": "4472C4",
    "設計": "5B9BD5",
    "開発": "70AD47",
    "Unit Test": "FFC000",
    "結合テスト": "ED7D31",
    "外部結合テスト": "E06020",
    "Non Functional Test": "BF8F00",
    "シナリオテスト": "C55A11",
    "UAT": "C00000",
    "リリース": "7030A0",
    # Hokkaido 用
    "PJ管理": "4472C4",
    "戦略定義": "5B9BD5",
    "スコープ定義": "70AD47",
    "構造定義": "FFC000",
    "スケルトン定義": "ED7D31",
    "UI/UX": "BF8F00",
    "骨格定義": "ED7D31",
    "表層定義": "BF8F00",
    "外部結合": "E06020",
    "テスト": "C55A11",
    "運用定義": "808080",
    "運用": "808080",
}

DEFAULT_COLOR = "808080"

# 工程ソート順
PROCESS_ORDER: list[str] = [
    "要件定義", "設計", "開発",
    "Unit Test", "結合テスト", "外部結合テスト",
    "Non Functional Test", "シナリオテスト", "UAT",
    "リリース",
    "PJ管理", "戦略定義", "スコープ定義", "構造定義",
    "スケルトン定義", "骨格定義", "表層定義", "UI/UX",
    "外部結合", "テスト", "運用定義", "運用",
]

# 固定列（A〜E）
INFO_COLUMNS = ["フェーズ", "工程", "開始日", "終了日", "タスク数"]
INFO_COL_COUNT = len(INFO_COLUMNS)


# ---------------------------------------------------------------------------
# データクラス
# ---------------------------------------------------------------------------

@dataclass
class GanttBar:
    group: str
    process: str
    label: str
    start: date
    end: date
    task_count: int
    is_milestone: bool = False  # 未使用（将来用に残置）


# ---------------------------------------------------------------------------
# 列名自動検出
# ---------------------------------------------------------------------------

def detect_columns(headers: list[str]) -> dict[str, str]:
    """CSVヘッダーからセマンティック列名マッピングを返す。"""
    col_map: dict[str, str] = {}
    header_set = set(headers)
    for semantic, candidates in COLUMN_CANDIDATES.items():
        for candidate in candidates:
            if candidate in header_set:
                col_map[semantic] = candidate
                break
    missing = {"start_date", "end_date", "group"} - col_map.keys()
    if missing:
        raise ValueError(
            f"必須列が見つかりません: {missing}。"
            f"CSVヘッダー: {headers[:15]}..."
        )
    return col_map


# ---------------------------------------------------------------------------
# CSV 読み込み・集約
# ---------------------------------------------------------------------------

def read_wbs_csv(
    path: Path, encoding: str = "utf-8-sig"
) -> tuple[list[dict[str, str]], list[str]]:
    """WBS CSV を読み込み、行辞書のリストとヘッダーを返す。"""
    with path.open("r", encoding=encoding, newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=[",", "\t", ";"])
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(f, dialect=dialect)
        headers = reader.fieldnames or []
        rows = list(reader)
    return rows, list(headers)


def _group_sort_key(group_name: str) -> tuple[int, str]:
    name = group_name.strip()
    if name.startswith("P"):
        return (0, name)
    if name.startswith("G"):
        return (0, name)
    return (1, name)


def _process_sort_key(process_name: str) -> int:
    try:
        return PROCESS_ORDER.index(process_name)
    except ValueError:
        return len(PROCESS_ORDER)


def aggregate_tasks(
    rows: list[dict[str, str]], col_map: dict[str, str]
) -> list[GanttBar]:
    """個別タスクをグループ×工程に集約し、GanttBar リストを返す。"""
    start_col = col_map["start_date"]
    end_col = col_map["end_date"]
    group_col = col_map["group"]
    process_col = col_map.get("subgroup", group_col)

    groups: dict[tuple[str, str], list[tuple[date, date]]] = defaultdict(list)
    skipped = 0

    for row in rows:
        group_val = (row.get(group_col) or "").strip()
        process_val = (row.get(process_col) or "").strip()
        start_str = (row.get(start_col) or "").strip()
        end_str = (row.get(end_col) or "").strip()

        if not group_val or not start_str or not end_str:
            skipped += 1
            continue

        # マイルストーン・リリース行はガントチャートから除外
        if process_val in ("マイルストーン", "リリース"):
            skipped += 1
            continue

        try:
            s = date.fromisoformat(start_str)
            e = date.fromisoformat(end_str)
        except ValueError:
            skipped += 1
            continue

        key = (group_val, process_val or group_val)
        groups[key].append((s, e))

    if skipped > 0:
        print(f"  info: {skipped} 行をスキップ（日付欠損等）")

    if not groups:
        raise ValueError("ガントチャートを生成するタスクがありません。")

    bars: list[GanttBar] = []
    sorted_keys = sorted(
        groups.keys(),
        key=lambda k: (_group_sort_key(k[0]), _process_sort_key(k[1])),
    )

    for group, process in sorted_keys:
        date_pairs = groups[(group, process)]
        starts = [d[0] for d in date_pairs]
        ends = [d[1] for d in date_pairs]
        earliest = min(starts)
        latest = max(ends)
        is_ms = (latest - earliest).days <= 1

        bars.append(
            GanttBar(
                group=group,
                process=process,
                label=f"{group}  {process}",
                start=earliest,
                end=latest,
                task_count=len(date_pairs),
                is_milestone=is_ms,
            )
        )

    return bars


# ---------------------------------------------------------------------------
# 週の月曜日リスト生成
# ---------------------------------------------------------------------------

def _monday_of(d: date) -> date:
    """d を含む週の月曜日を返す。"""
    return d - timedelta(days=d.weekday())


def generate_week_columns(bars: list[GanttBar]) -> list[date]:
    """全バーの範囲をカバーする週（月曜始まり）のリストを返す。"""
    all_starts = [b.start for b in bars]
    all_ends = [b.end for b in bars]
    first_monday = _monday_of(min(all_starts))
    last_monday = _monday_of(max(all_ends))
    weeks: list[date] = []
    current = first_monday
    while current <= last_monday:
        weeks.append(current)
        current += timedelta(days=7)
    return weeks


# ---------------------------------------------------------------------------
# Excel ガントチャート書き込み
# ---------------------------------------------------------------------------

def _font_color_for_bg(hex_color: str) -> str:
    """背景色に対するフォント色（白 or 黒）を返す。"""
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    luminance = 0.299 * r + 0.587 * g + 0.114 * b
    return "FFFFFF" if luminance < 160 else "000000"


def write_gantt_xlsx(
    bars: list[GanttBar], weeks: list[date], output_xlsx: Path, title: str
) -> None:
    """ガントチャートを Excel ファイルに書き込む。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "ガントチャート"

    thin = Side(style="thin", color="D0D7DE")
    hair = Side(style="hair", color="E0E0E0")
    border_info = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_week = Border(left=hair, right=hair, top=thin, bottom=thin)

    header_font = Font(name="Yu Gothic", bold=True, color="FFFFFF", size=9)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    info_font = Font(name="Yu Gothic", size=9)
    info_align = Alignment(vertical="center", wrap_text=False)

    month_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    month_font = Font(name="Yu Gothic", bold=True, size=8)

    today = date.today()
    today_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    group_separator_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # --- 行1: 月ヘッダー（結合セル） ---
    row1 = 1
    # 情報列の結合ヘッダー
    ws.merge_cells(start_row=row1, start_column=1, end_row=row1, end_column=INFO_COL_COUNT)
    cell = ws.cell(row=row1, column=1, value=title)
    cell.font = Font(name="Yu Gothic", bold=True, color="FFFFFF", size=10)
    cell.fill = header_fill
    cell.alignment = header_align
    for c in range(2, INFO_COL_COUNT + 1):
        ws.cell(row=row1, column=c).fill = header_fill

    # 月ごとに結合
    month_ranges: list[tuple[int, int, str]] = []
    prev_ym = ""
    start_col_idx = INFO_COL_COUNT + 1
    for i, monday in enumerate(weeks):
        col_idx = INFO_COL_COUNT + 1 + i
        ym = monday.strftime("%Y-%m")
        if ym != prev_ym:
            if month_ranges:
                month_ranges[-1] = (month_ranges[-1][0], col_idx - 1, month_ranges[-1][2])
            month_ranges.append((col_idx, col_idx, ym))
            prev_ym = ym
        else:
            month_ranges[-1] = (month_ranges[-1][0], col_idx, ym)

    for start_c, end_c, ym in month_ranges:
        if end_c > start_c:
            ws.merge_cells(start_row=row1, start_column=start_c, end_row=row1, end_column=end_c)
        cell = ws.cell(row=row1, column=start_c, value=ym)
        cell.font = month_font
        cell.fill = month_fill
        cell.alignment = header_align
        cell.border = border_info
        for c in range(start_c + 1, end_c + 1):
            ws.cell(row=row1, column=c).fill = month_fill
            ws.cell(row=row1, column=c).border = border_info

    # --- 行2: 列ヘッダー（情報列 + 週列） ---
    row2 = 2
    for i, col_name in enumerate(INFO_COLUMNS, start=1):
        cell = ws.cell(row=row2, column=i, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border_info

    for i, monday in enumerate(weeks):
        col_idx = INFO_COL_COUNT + 1 + i
        label = f"{monday.month}/{monday.day}"
        cell = ws.cell(row=row2, column=col_idx, value=label)
        cell.font = Font(name="Yu Gothic", size=7, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border_info

    # --- データ行 ---
    prev_group = None
    for row_i, bar in enumerate(bars):
        r = row_i + 3  # データは行3から

        # フェーズ境界で空行風に薄い背景を適用
        is_group_change = prev_group is not None and bar.group != prev_group
        prev_group = bar.group

        # 情報列
        info_values = [
            bar.group,
            bar.process,
            bar.start.isoformat(),
            bar.end.isoformat(),
            bar.task_count,
        ]
        for c, val in enumerate(info_values, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = info_font
            cell.alignment = info_align
            cell.border = border_info

        # 週列にバーを描画
        color_hex = PROCESS_COLORS.get(bar.process, DEFAULT_COLOR)
        bar_fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        font_hex = _font_color_for_bg(color_hex)
        for i, monday in enumerate(weeks):
            col_idx = INFO_COL_COUNT + 1 + i
            week_end = monday + timedelta(days=6)
            cell = ws.cell(row=r, column=col_idx)
            cell.border = border_week

            # 今日の週をハイライト
            is_today_week = monday <= today <= week_end

            # バー範囲内か判定
            if bar.start <= week_end and bar.end >= monday:
                cell.fill = bar_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif is_today_week:
                cell.fill = today_fill

    # --- 列幅 ---
    ws.column_dimensions["A"].width = 8   # フェーズ
    ws.column_dimensions["B"].width = 18  # 工程
    ws.column_dimensions["C"].width = 12  # 開始日
    ws.column_dimensions["D"].width = 12  # 終了日
    ws.column_dimensions["E"].width = 6   # タスク数

    for i in range(len(weeks)):
        col_letter = get_column_letter(INFO_COL_COUNT + 1 + i)
        ws.column_dimensions[col_letter].width = 4.5

    # 行高
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    for r in range(3, 3 + len(bars)):
        ws.row_dimensions[r].height = 20

    # フリーズ
    ws.freeze_panes = ws.cell(row=3, column=INFO_COL_COUNT + 1).coordinate

    # --- 凡例シート ---
    ws_legend = wb.create_sheet("凡例")
    ws_legend.cell(row=1, column=1, value="工程").font = Font(bold=True)
    ws_legend.cell(row=1, column=2, value="色").font = Font(bold=True)
    used = sorted({b.process for b in bars}, key=_process_sort_key)
    for i, proc in enumerate(used, start=2):
        ws_legend.cell(row=i, column=1, value=proc)
        color = PROCESS_COLORS.get(proc, DEFAULT_COLOR)
        cell = ws_legend.cell(row=i, column=2, value="")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    ws_legend.column_dimensions["A"].width = 24
    ws_legend.column_dimensions["B"].width = 12

    # 保存
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="WBS CSV からガントチャート Excel を生成します。"
    )
    parser.add_argument("input_csv", type=Path, help="入力 WBS CSV ファイル")
    parser.add_argument(
        "-o", "--output", type=Path,
        help="出力 Excel ファイル（省略時は入力と同名 _gantt.xlsx）",
    )
    parser.add_argument(
        "--to-deliverable", action="store_true",
        help="`docs/exports/` へ出力する",
    )
    parser.add_argument(
        "--project-root", type=Path,
        help="`--to-deliverable` 時のPJルート（省略時は入力ファイルから推定）",
    )
    parser.add_argument(
        "--date", dest="deliverable_date",
        help="`--to-deliverable` 時の日付（YYYY-MM-DD, 省略時は当日）",
    )
    parser.add_argument(
        "--encoding", default="utf-8-sig",
        help="CSV文字コード（既定: utf-8-sig）",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="変換は実行せず、出力先のみ表示する",
    )
    return parser.parse_args()


def infer_project_root(input_path: Path) -> Path | None:
    required = {"0_Project", "9_Decision_Records"}
    for parent in [input_path.parent, *input_path.parents]:
        try:
            dir_names = {child.name for child in parent.iterdir() if child.is_dir()}
        except OSError:
            continue
        if required.issubset(dir_names):
            return parent
    return None


def resolve_output_xlsx(
    args: argparse.Namespace, input_csv: Path, *, dry_run: bool = False
) -> Path:
    if args.output:
        output = args.output.resolve()
        if not dry_run:
            output.parent.mkdir(parents=True, exist_ok=True)
        return output

    if not args.to_deliverable:
        return input_csv.with_name(f"{input_csv.stem}_gantt.xlsx")

    if args.deliverable_date:
        try:
            export_date = date.fromisoformat(args.deliverable_date).isoformat()
        except ValueError as exc:
            raise ValueError(
                f"--date は YYYY-MM-DD 形式で指定してください: {args.deliverable_date}"
            ) from exc
    else:
        export_date = date.today().isoformat()

    project_root = (
        args.project_root.resolve() if args.project_root
        else infer_project_root(input_csv)
    )
    if project_root is None:
        raise ValueError(
            "--to-deliverable 利用時にPJルートを推定できませんでした。"
            " `--project-root <path>` を指定してください。"
        )

    deliverable_dir = project_root / "docs" / "exports"
    if not dry_run:
        deliverable_dir.mkdir(parents=True, exist_ok=True)
    return deliverable_dir / f"{input_csv.stem}_gantt.xlsx"


def main() -> int:
    args = parse_args()
    input_csv = args.input_csv.resolve()

    if not input_csv.exists():
        print(f"ERROR: 入力ファイルが見つかりません: {input_csv}", file=sys.stderr)
        return 1

    try:
        output_xlsx = resolve_output_xlsx(args, input_csv, dry_run=args.dry_run)
    except ValueError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    if args.dry_run:
        print("INFO: Dry run mode: 変換は実行しません。")
        print(f"   input : {input_csv}")
        print(f"   output: {output_xlsx}")
        return 0

    try:
        rows, headers = read_wbs_csv(input_csv, args.encoding)
    except Exception as exc:
        print(f"ERROR: CSV読み込みに失敗しました: {exc}", file=sys.stderr)
        return 1

    try:
        col_map = detect_columns(headers)
    except ValueError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    print(f"  columns: {col_map}")

    try:
        bars = aggregate_tasks(rows, col_map)
    except ValueError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    weeks = generate_week_columns(bars)
    print(f"  bars: {len(bars)} 集約バー（{len(rows)} 行から生成）, {len(weeks)} 週列")

    title = f"{input_csv.stem} ガントチャート"

    try:
        write_gantt_xlsx(bars, weeks, output_xlsx, title)
    except Exception as exc:
        print(f"ERROR: ガントチャート書き込みに失敗しました: {exc}", file=sys.stderr)
        return 1

    print(f"OK: ガントチャート Excel 出力完了: {output_xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
